#!/usr/bin/env python3
"""JSON -> XLSX. Input: {"SheetName": [ {col: val, ...}, ... ], ...}.
One sheet per top-level key, header row from union of keys (first-seen order),
auto-width, frozen header. Usage:
    make_xlsx.py --out report.xlsx --json data.json    # or --json - for stdin
"""
import argparse, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def col_order(rows):
    seen = {}
    for r in rows:
        for k in r:
            seen.setdefault(k, None)
    return list(seen)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--json", required=True, help="path to JSON, or - for stdin")
    a = ap.parse_args()

    raw = sys.stdin.read() if a.json == "-" else open(a.json, encoding="utf-8").read()
    data = json.loads(raw)
    if not isinstance(data, dict):
        sys.exit("JSON must be an object {sheet: [rows...]}")

    wb = Workbook()
    wb.remove(wb.active)
    for sheet, rows in data.items():
        ws = wb.create_sheet(title=str(sheet)[:31])  # Excel 31-char limit
        rows = rows or []
        cols = col_order(rows)
        if not cols:
            ws.append(["(no data)"])
            continue
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(k)) for k in cols])
        ws.freeze_panes = "A2"
        for i, k in enumerate(cols, 1):
            width = max(len(str(k)), *(len(str(_cell(r.get(k)))) for r in rows))
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)
    if not wb.sheetnames:
        wb.create_sheet("Empty")
    wb.save(a.out)
    print(f"wrote {a.out}")


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


if __name__ == "__main__":
    main()
